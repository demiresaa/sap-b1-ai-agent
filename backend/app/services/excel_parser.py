"""Excel sipariş listesi parser'ı.

openpyxl ile ilk sheet'i okur, header satırını (ilk dolu satır) ve veri satırlarını
ayırır. AI'ya gönderilmek üzere "tabular" bir dict üretir — kolon adlarını tahmin
etmeyiz, AI mapping yapar.

Multi-sheet workbook'larda şu an sadece ilk dolu sheet kullanılır.
"""
from __future__ import annotations

import logging
import pathlib
from dataclasses import dataclass, field
from typing import Any

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# Tüm hücreler boş ise satırı iskele (header probe).
MAX_HEADER_PROBE_ROWS = 5


@dataclass(slots=True)
class ExcelSheet:
    """Tek bir sheet'in normalize edilmiş hâli."""

    name: str
    headers: list[str]
    rows: list[dict[str, Any]] = field(default_factory=list)
    raw_rows_count: int = 0

    def to_prompt_text(self, max_rows: int = 50) -> str:
        """LLM prompt'una gönderilebilir kompakt TSV temsili."""
        header_line = "\t".join(self.headers)
        sample = self.rows[:max_rows]
        body_lines = [
            "\t".join(str(r.get(h, "") if r.get(h) is not None else "") for h in self.headers)
            for r in sample
        ]
        suffix = ""
        if len(self.rows) > max_rows:
            suffix = f"\n… ve {len(self.rows) - max_rows} satır daha"
        return f"Sheet: {self.name}\n{header_line}\n" + "\n".join(body_lines) + suffix


def parse_excel(path: pathlib.Path | str) -> ExcelSheet:
    """İlk dolu sheet'i okuyup `ExcelSheet` döner.

    Heuristik header tespiti: İlk 5 satırda dolu hücre sayısı en yüksek olan satır
    header kabul edilir; sonraki tüm satırlar veri.
    """
    p = pathlib.Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Excel bulunamadı: {p}")

    wb = load_workbook(p, data_only=True, read_only=True)
    target_ws = None
    for ws in wb.worksheets:
        if ws.max_row and ws.max_row > 0:
            target_ws = ws
            break
    if target_ws is None:
        raise ValueError("Excel boş — sheet bulunamadı.")

    # Tüm satırları kapla (header probe için ilk N satıra bak)
    all_rows: list[list[Any]] = []
    for row in target_ws.iter_rows(values_only=True):
        all_rows.append(list(row))
    if not all_rows:
        return ExcelSheet(name=target_ws.title, headers=[], rows=[], raw_rows_count=0)

    header_idx = _detect_header_row(all_rows)
    raw_headers = all_rows[header_idx]
    headers = _normalize_headers(raw_headers)

    rows: list[dict[str, Any]] = []
    for raw in all_rows[header_idx + 1 :]:
        if all(cell is None or (isinstance(cell, str) and not cell.strip()) for cell in raw):
            continue
        row_dict: dict[str, Any] = {}
        for i, header in enumerate(headers):
            value = raw[i] if i < len(raw) else None
            row_dict[header] = _normalize_cell(value)
        rows.append(row_dict)

    logger.info(
        "[excel] sheet=%s header_idx=%d cols=%d rows=%d",
        target_ws.title,
        header_idx,
        len(headers),
        len(rows),
    )
    return ExcelSheet(
        name=target_ws.title,
        headers=headers,
        rows=rows,
        raw_rows_count=len(all_rows),
    )


def _detect_header_row(rows: list[list[Any]]) -> int:
    """İlk birkaç satırda en fazla dolu hücre olan satırı header kabul et."""
    best_idx = 0
    best_filled = -1
    for i, row in enumerate(rows[:MAX_HEADER_PROBE_ROWS]):
        filled = sum(
            1 for c in row if c is not None and not (isinstance(c, str) and not c.strip())
        )
        if filled > best_filled:
            best_filled = filled
            best_idx = i
    return best_idx


def _normalize_headers(raw: list[Any]) -> list[str]:
    seen: dict[str, int] = {}
    out: list[str] = []
    for i, h in enumerate(raw):
        name = str(h).strip() if h is not None else ""
        if not name:
            name = f"col_{i + 1}"
        # Duplicate kolon adlarına suffix
        if name in seen:
            seen[name] += 1
            name = f"{name}__{seen[name]}"
        else:
            seen[name] = 1
        out.append(name)
    return out


def _normalize_cell(value: Any) -> Any:
    if isinstance(value, str):
        v = value.strip()
        return v if v else None
    return value
